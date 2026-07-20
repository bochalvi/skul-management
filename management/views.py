from decimal import Decimal
from multiprocessing import context

from django.shortcuts import render
from django.db.models import Sum, F, Avg
from django.db import transaction
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart,PieChart, Reference

from django.http import HttpResponse


# Create your views here.
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, JsonResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from datetime import date,datetime, timedelta
import csv
from.models import *
from .forms import *

# Create your views here.

def is_teacher(user):
    return user.groups.filter(name='Teachers').exists()

def is_admin(user):
    return user.is_superuser

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'management/home.html')

def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Registration successful. You can now log in.')
            return redirect('login')
    else:
        form = UserRegistrationForm()
    return render(request, 'management/register.html', {'form': form})

@login_required
def dashboard(request):
    # get statistics 
    total_students = Student.objects.filter(is_active=True).count()
    
    today = date.today()
    
    # Calculate today's attendance percentage
    today_attendance = Attendance.objects.filter(attendance_date=today)
    total_present = today_attendance.filter(status='Present').count()
    total_students_count = Student.objects.filter(is_active=True).count()
    present_today = round((total_present / total_students_count * 100), 1) if total_students_count > 0 else 0
    
    # Calculate pending fees
    pending_fees = Fee.objects.filter(status__in=['Unpaid', 'partially paid']).aggregate(
        total=Sum('amount_due') - Sum('amount_paid')
    )['total'] or 0
    
    today_birthdays = Student.objects.filter(
        date_of_birth__month=today.month, 
        date_of_birth__day=today.day
    )
    today_lessons = LessonPlan.objects.filter(date=today, status='Planned').count()
    
    # Recent activities - Fixed: removed 'class_attended'
    recent_attendance = Attendance.objects.select_related('student').order_by('-attendance_date')[:5]
    recent_fees = Fee.objects.select_related('student', 'fee_type').order_by('-due_date')[:5]
    
    context = {
        'total_students': total_students,
        'present_today': present_today,
        'pending_fees': pending_fees,
        'today_birthdays': today_birthdays,
        'today_lessons': today_lessons,
        'recent_attendance': recent_attendance,
        'recent_fees': recent_fees,
    }
    return render(request, 'management/dashboard.html', context)

# Student views
class StudentListView(LoginRequiredMixin, ListView):
    model = Student
    template_name = 'management/student_list.html'
    context_object_name = 'students'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(First_Name__icontains=search_query) | 
                Q(Last_Name__icontains=search_query) | 
                Q(Admission_Number__icontains=search_query)|
                Q(Parent_Guardian_Name__icontains=search_query)
        )
        active_filter = self.request.GET.get('active', '')
        if active_filter == 'active':
            queryset = queryset.filter(is_active=True)
        elif active_filter == 'inactive':
            queryset = queryset.filter(is_active=False)
        return queryset.order_by('Last_Name', 'First_Name', 'Admission_Number', 'Enrollment_Date')
    
class StudentCreateView(LoginRequiredMixin, CreateView):
    model = Student
    form_class = StudentRegistrationForm
    template_name = 'management/student_form.html'
    success_url = reverse_lazy('student_list')
    
    
    
    
class StudentUpdateView(LoginRequiredMixin, UpdateView):
    model = Student
    form_class = StudentRegistrationForm
    template_name = 'management/student_form.html'
    success_url = reverse_lazy('student_list')
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
class StudentDeleteView(LoginRequiredMixin, DeleteView):
    model = Student
    success_url = reverse_lazy('student_list')
    template_name =  'management/student_confirm_delete.html'
    
    
class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student
    template_name = 'management/student_detail.html'
    context_object_name = 'student'
    
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
    
        # Get attendance records for the student - Fixed: use 'student' not 'student_id'
        attendance_records = Attendance.objects.filter(student=student).order_by('-attendance_date')
    
        # Get fee records for the student
        fee_records = Fee.objects.filter(student=student).order_by('-due_date')[:10]
    
        # Calculate statistics
        total_fees_due = fee_records.aggregate(total=Sum('amount_due'))['total'] or 0
        total_fees_paid = fee_records.aggregate(total=Sum('amount_paid'))['total'] or 0
        total_fees_pending = total_fees_due - total_fees_paid  # ✅ Fixed calculation
    
        # Calculate attendance rate
        total_attendance_days = attendance_records.count()
        present_days = attendance_records.filter(status='Present').count()
        attendance_rate = round((present_days / total_attendance_days * 100), 1) if total_attendance_days > 0 else 0
    
        # Get recent grades for the student
        grades = Grade.objects.filter(student=student).order_by('-recorded_at')[:10]
    
        context.update({
            'attendance_records': attendance_records,
            'fee_records': fee_records,
            'total_fees_due': total_fees_due,
            'total_fees_paid': total_fees_paid,
            'total_fees_pending': total_fees_pending,
            'attendance_rate': attendance_rate,  # ✅ Added attendance_rate
            'grades': grades,
        })
        return context
    
# Attendance views
@login_required
def attendance_list(request):
    selected_date = request.GET.get('date', date.today().isoformat())
    
    try:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    except:
        selected_date = date.today()
    
    # Get attendance records for the selected date
    attendance_records = Attendance.objects.filter(attendance_date=selected_date)
    
    # Get all active students
    all_students = Student.objects.filter(is_active=True)
    
    # Create dictionary mapping student ID to attendance record
    attendance_dict = {att.student.Admission_Number: att for att in attendance_records}
    
    if request.method == 'POST':
        for student in all_students:
            status = request.POST.get(f'status_{student.Admission_Number}', '')
            if status:
                check_in = request.POST.get(f'check_in_{student.Admission_Number}', '')
                check_out = request.POST.get(f'check_out_{student.Admission_Number}', '')
                notes = request.POST.get(f'notes_{student.Admission_Number}', '')
                
                attendance, created = Attendance.objects.update_or_create(
                    student=student,
                    attendance_date=selected_date,
                    defaults={
                        'status': status,
                        'check_in_time': check_in if check_in else None,
                        'check_out_time': check_out if check_out else None,
                        'notes': notes,
                        'recorded_by': request.user,
                    }
                )
        messages.success(request, f'Attendance for {selected_date} saved successfully!')
        return redirect('attendance_list')
    
    context = {
        'attendance_records': attendance_records,
        'attendance_dict': attendance_dict,
        'all_students': all_students,
        'selected_date': selected_date,
    }
    return render(request, 'management/attendance_list.html', context)

@login_required
def attendance_report(request):
    from datetime import datetime, date, timedelta
    from django.db.models import Count, Q
    
    # Get date parameters from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Set default dates (last 30 days if no dates provided)
    if not start_date_str:
        start_date = date.today() - timedelta(days=30)
    else:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = date.today() - timedelta(days=30)
    
    if not end_date_str:
        end_date = date.today()
    else:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = date.today()
    
    # Calculate total days in range
    total_days_in_range = (end_date - start_date).days + 1
    
    # Get all active students
    students = Student.objects.filter(is_active=True)
    
    # Get attendance records within date range
    attendance_records = Attendance.objects.filter(
        attendance_date__range=[start_date, end_date]
    )
    
    # Calculate statistics per student
    attendance_data = []
    total_present_all = 0
    total_absent_all = 0
    total_late_all = 0
    total_excused_all = 0
    
    for student in students:
        # Get records for this student
        student_records = attendance_records.filter(student=student)
        total_days = student_records.count()
        
        present_days = student_records.filter(status='Present').count()
        absent_days = student_records.filter(status='Absent').count()
        late_days = student_records.filter(status='Late').count()
        excused_days = student_records.filter(status='Excused').count()
        
        # Calculate attendance percentage
        if total_days > 0:
            attendance_percentage = round((present_days / total_days) * 100, 1)
        else:
            attendance_percentage = 0
        
        attendance_data.append({
            'student_first_name': student.First_Name,
            'student_last_name': student.Last_Name,
            'student_admission_number': student.Admission_Number,
            'total_days': total_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'excused_days': excused_days,
            'attendance_percentage': attendance_percentage,
        })
        
        total_present_all += present_days
        total_absent_all += absent_days
        total_late_all += late_days
        total_excused_all += excused_days
    
    # Calculate overall attendance percentage
    total_present_absent = total_present_all + total_absent_all
    if total_present_absent > 0:
        overall_attendance = round((total_present_all / total_present_absent) * 100, 1)
    else:
        overall_attendance = 0
    
    context = {
        'attendance_data': attendance_data,
        'start_date': start_date,
        'end_date': end_date,
        'total_days': total_days_in_range,
        'overall_attendance': overall_attendance,
        'total_present': total_present_all,
        'total_absent': total_absent_all,
        'total_late': total_late_all,
        'total_excused': total_excused_all,
    }
    
    return render(request, 'management/attendance_report.html', context)
# Fee views
class FeeListView(LoginRequiredMixin, ListView):
    model = Fee
    template_name = 'management/fee_list.html'
    context_object_name = 'fees'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Fee.objects.all().select_related('student','fee_type')
        # filter by student name or admission number
        status_filter = self.request.GET.get('status')
        if status_filter :
            queryset = queryset.filter(status=status_filter)
            
        student_filter = self.request.GET.get('student')
        if student_filter:
            queryset = queryset.filter(student__Admission_Number=student_filter)
            
        return queryset.order_by('-due_date', 'student__Last_Name', 'student__First_Name')
    
    def get_context_data(self, **kwargs):
        context =super().get_context_data(**kwargs)
        # total due for pending/partial fees
        total_due_result = Fee.objects.filter(status__in=['Unpaid','partially paid']).aggregate(
            total=Sum(F('amount_due') - F('amount_paid')))
        total_due = total_due_result['total'] if total_due_result['total']  else 0
        context['total_due'] = total_due
        # total collected from paid fees
        total_collected = Fee.objects.filter(status='Paid').aggregate(
            total=Sum('amount_paid'))['total'] or 0
        context['total_collected'] = total_collected
        # overdue count
        from datetime import date
        context['overdue_fees'] = Fee.objects.filter(
            due_date__lt=date.today(),
            status__in=['Unpaid', 'partially paid']
        ).count()
        context['students'] = Student.objects.filter(is_active=True).order_by('Last_Name', 'First_Name')
        return context
    
class FeeCreateView(LoginRequiredMixin, CreateView):
    model = Fee
    form_class = FeeForm
    template_name = 'management/fee_form.html'
    success_url = reverse_lazy('fee_list')
    
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # update status based on payment
        if form.instance.amount_paid >= form.instance.amount_due:
            form.instance.status = 'Paid'
            if not form.instance.paid_date:
                form.instance.paid_date = date.today()
            elif form.instance.amount_paid > 0:
                form.instance.status = 'Partially Paid'
            else:
                form.instance.status = 'Pending'
            messages.success(self.request, 'Fee record created successfully.')
        return super().form_valid(form)
    
@login_required
def fee_payment(request, pk):
    fee = get_object_or_404(Fee, pk=pk)
    
    if request.method == 'POST':
        amount_paid_str = request.POST.get('amount_paid')
        payment_method = request.POST.get('payment_method')
        transaction_id = request.POST.get('transaction_id')
        
        try:
            amount_paid = Decimal(amount_paid_str)
            
            # Check if amount_paid exceeds balance
            balance = fee.amount_due - fee.amount_paid
            if amount_paid > balance:
                messages.error(request, f'Payment amount cannot exceed the balance of Kshs {balance:.2f}')
                return redirect('fee_payment', pk=fee.pk)
            
            # Update fee record
            fee.amount_paid += amount_paid
            fee.payment_method = payment_method
            fee.transaction_id = transaction_id
            fee.paid_date = date.today()
            
            # Update status based on payment
            if fee.amount_paid >= fee.amount_due:
                fee.status = 'Paid'
            elif fee.amount_paid > 0:
                fee.status = 'partially paid'
            else:
                fee.status = 'Unpaid'
            
            fee.save()
            messages.success(request, f'Payment of Kshs {amount_paid:.2f} recorded successfully.')
            return redirect('fee_list')
            
        except ValueError:
            messages.error(request, 'Invalid amount. Please enter a valid number.')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
    
    # Calculate balance for display
    balance = fee.amount_due - fee.amount_paid
    
    context = {
        'fee': fee,
        'balance': balance,
    }
    return render(request, 'management/fee_payment.html', context)

class FeeDeleteView(LoginRequiredMixin,DeleteView):
    model = Fee
    success_url = reverse_lazy('fee_list')
    context_object_name = 'fee'
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.groups.filter(name='Admin').exist()
    
@login_required
def generate_fees(request):
    if request.method == 'POST':
        month = request.POST.get('month')
        year = request.POST.get('year')
        fee_type_id = request.POST.get('fee_type')

        if not (month and year and fee_type_id):
            messages.error(request, 'Please select month, year, and fee type.')
            return redirect('generate_fees')

        try:
            month = int(month)
            year = int(year)
            fee_type = FeeType.objects.get(id=fee_type_id)
        except (ValueError, FeeType.DoesNotExist):
            messages.error(request, 'Invalid selection.')
            return redirect('generate_fees')

        due_date = date(year, month, 10)
        students = Student.objects.filter(is_active=True)

        created_count = 0
        skipped_count = 0

        with transaction.atomic():
            for student in students:
                existing = Fee.objects.filter(
                    student=student,
                    fee_type=fee_type,
                    due_date__year=year,
                    due_date__month=month
                ).exists()

                if not existing:
                    Fee.objects.create(
                        student=student,
                        fee_type=fee_type,
                        amount_due=fee_type.amount,
                        amount_paid=0,
                        due_date=due_date,
                        status='Unpaid',
                        notes=f'Generated for {month}/{year}'
                    )
                    created_count += 1
                else:
                    skipped_count += 1

        messages.success(request, f'Fees generated: {created_count} created, {skipped_count} already existed.')
        return redirect('fee_list')

    # GET request
    fee_types = FeeType.objects.filter(is_active=True)
    current_year = date.today().year
    years = range(current_year - 1, current_year + 2)
    months = range(1, 13)

    context = {
        'fee_types': fee_types,
        'years': years,
        'months': months,
        'current_year': current_year,
        'current_month': date.today().month,
    }
    return render(request, 'management/generate_fees.html', context)

class LessonPlanListView(LoginRequiredMixin, ListView):
    model = LessonPlan
    template_name = 'management/lessonPlan_list.html'
    context_object_name = 'lessonplans'  # Fixed: should match template variable
    paginate_by = 20
    
    def get_queryset(self):
        queryset = LessonPlan.objects.all().order_by('-date')
        if self.request.user.groups.filter(name='Teachers').exists():
            queryset = queryset.filter(teacher=self.request.user)
        date_filter = self.request.GET.get('date', '')
        if date_filter:
            try:
                date_filter = datetime.strptime(date_filter, '%Y-%m-%d').date()
                queryset = queryset.filter(date=date_filter)
            except:
                pass
            
        status_filter = self.request.GET.get('status', '')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        subject_filter = self.request.GET.get('subject', '')
        if subject_filter:
            queryset = queryset.filter(subject__icontains=subject_filter)
            
        return queryset

# ✅ FIXED: This class should NOT be indented inside LessonPlanListView
class LessonPlanCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = LessonPlan
    form_class = LessonPlanForm
    template_name = 'management/lessonPlan_form.html'
    success_url = reverse_lazy('lesson_plan_list')
    
    def form_valid(self, form):
        form.instance.teacher = self.request.user
        messages.success(self.request, 'Lesson plan created successfully.')
        return super().form_valid(form)
    
    def test_func(self):
        # Allow only staff or teachers to create lesson plans
        return self.request.user.is_staff or self.request.user.groups.filter(name='Teacher').exists()
        
class LessonPlanUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = LessonPlan
    form_class = LessonPlanForm
    template_name = 'management/lessonPlan_form.html'
    success_url = reverse_lazy('lesson_plan_list')
    
    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.groups.filter(name='Teachers').exists():
            queryset = queryset.filter(teacher=self.request.user)
        return queryset
    
    def test_func(self):
        # Allow only the teacher who created it, staff, or admins to edit
        lessonplan = self.get_object()
        return (self.request.user.is_staff or 
                self.request.user.groups.filter(name='Teacher').exists() or
                lessonplan.teacher == self.request.user)
    
class LessonPlanDetailView(LoginRequiredMixin, DetailView):
    model = LessonPlan
    template_name = 'management/lessonPlan_detail.html'
    context_object_name = 'lessonplan'  # Fixed: should match template variable
    
class LessonPlanDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = LessonPlan
    template_name = 'management/lessonPlan_confirm_delete.html'
    success_url = reverse_lazy('lesson_plan_list')
    
    def test_func(self):
        lessonplan = self.get_object()
        return (self.request.user.is_staff or 
                self.request.user.groups.filter(name='Teacher').exists() or
                lessonplan.teacher == self.request.user)    
    
 # report generation  
@login_required
def generate_report(request):
    report_type = request.GET.get('type','attendance')
    
    if report_type == 'attendance':
        return attendance_report(request)
    elif report_type == 'fees':
        return fee_report(request)
    elif report_type == 'students':
        return student_report(request)
    
    return HttpResponse('Invalid report type', status=400)

@login_required
def student_report(request):
    """Generate a report of all students with their attendance and fee status."""
    students = Student.objects.filter(is_active=True).select_related('user')
    
    # Filter by grade/class if provided
    grade_filter = request.GET.get('grade')
    if grade_filter:
        students = students.filter(grade=grade_filter)
    
    # Get attendance summary for each student
    attendance_summary = {}
    for student in students:
        total_present = Attendance.objects.filter(student=student, status='present').count()
        total_absent = Attendance.objects.filter(student=student, status='absent').count()
        attendance_summary[student.id] = {
            'present': total_present,
            'absent': total_absent,
        }
    
    # Get fee summary for each student
    fee_summary = {}
    for student in students:
        total_due = Fee.objects.filter(student=student).aggregate(total=Sum('amount_due'))['total'] or 0
        total_paid = Fee.objects.filter(student=student).aggregate(total=Sum('amount_paid'))['total'] or 0
        fee_summary[student.id] = {
            'total_due': total_due,
            'total_paid': total_paid,
            'balance': total_due - total_paid,
        }
    
    context = {
        'students': students,
        'attendance_summary': attendance_summary,
        'fee_summary': fee_summary,
    }
    return render(request, 'management/student_report.html', context)

@login_required
def fee_report(request):
    start_date = request.GET.get('start_date',(date.today() - timedelta(days=90)).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except:
        start_date = date.today() - timedelta(days=90)
        end_date = date.today()
        
    fee_data = Fee.objects.filter(
        created_at__date__range=[start_date, end_date]
    ).values('fee_type__name').annotate(
        total_due=Sum('amount_due'),
        total_paid=Sum('amount_paid'),
        count=Count('id')
    )
    
    
    payment_methods = Fee.objects.filter(
        created_at__date__range=[start_date, end_date],
        status='Paid'
    ).values('payment_method').annotate(
        total=Sum('amount_paid'),
        count=Count('id')
    )
    
    context = {
        'fee_data': fee_data,
        'payment_methods': payment_methods,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'management/fee_report.html', context)

# export data
@login_required
def export_students_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Admission Number', 'First Name', 'Last Name', 'Gender', 'Grade',
        'Parent_Guardian_Name','Parent_Guardian_Contact','Parent_Guardian_Email', 'Enrollment Date', 'Active'
    ])
    
    
    students = Student.objects.all()
    for student in students:
        writer.writerow([
            student.Admission_Number,
            student.First_Name,
            student.Last_Name,
            student.date_of_birth,
            student.get_Gender_display(),
            
            student.Parent_Guardian_Name,
            student.Parent_Guardian_Contact,
            student.Enrollment_Date,
            'Yes' if student.is_active else 'No'
        ])
    
    return response
 # export attendance data      
@login_required
def export_attendance_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Student', 'Date','check_in','check_out', 'Status'])
    
    start_date = request.GET.get('start_date', (date.today() - timedelta(days=30)).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    try:
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    except:
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()
    
    attendance_records = Attendance.objects.filter(
        attendance_date__range=[start_date, end_date]
    ).select_related('student')
    for record in attendance_records:
        writer.writerow([
            record.student.full_name,
            record.attendance_date,
            record.check_in_time,
            record.check_out_time,
            record.status
        ])
    
    return response  

@login_required
def export_fees_csv(request):
    queryset = Fee.objects.select_related('student','fee_type').all() 
    status_filter = request.GET.get('status')
    if status_filter:
        queryset = queryset.filter(status=status_filter)
        
    student_filter = request.GET.get('student')
    if student_filter and student_filter.isdigit():
        queryset = queryset.filter(student__Admission_Number=student_filter)
        
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            queryset = queryset.filter(due_date__gte=start_date)
        except ValueError:
            pass
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(due_date__lte=end_date)
        except ValueError:
            pass
        
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="fees_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Admission Number',
        'Student Name',
        'Fee Type',
        'Amount Due',
        'Amount Paid',
        'Balance',
        'Due Date',
        'Paid Date',
        'Status',
        'Payment Method',
        'Transaction ID',
        'Notes',
        'Created at'
    ])
    for fee in queryset:
        balance = fee.amount_due - fee.amount_paid
        writer.writerow([
            fee.student.Admission_Number,
            fee.student.full_name,
            fee.fee_type.name if fee.fee_type else '',
            str(fee.amount_due),
            str(fee.amount_paid),
            str(balance),
            fee.due_date.strftime('%Y-%m-%d')if fee.due_date else '',
            fee.paid_date.strftime('%Y-%m-%d')if fee.paid_date else '',
            fee.get_status_display(),
            fee.payment_method or '',
            fee.transaction_id or '',
            fee.notes or '',
            fee.created_at.strftime('%Y-%m-%d %H:%M:%S') if fee.created_at else '',
        ])
    return response
    
@login_required
def student_tutorial_list(request):
    """Show all published tutorials with progress for the logged-in student."""
    # Assuming the logged-in user is linked to a Student record
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        # If user is not a student (e.g., teacher), show all tutorials without progress
        student = None

    tutorials = Tutorial.objects.filter(is_published=True)

    if student:
        # Annotate each tutorial with completion status
        progress = StudentTutorialProgress.objects.filter(student=student)
        progress_dict = {p.tutorial_id: p.completed for p in progress}
        for tutorial in tutorials:
            tutorial.completed = progress_dict.get(tutorial.id, False)
    else:
        for tutorial in tutorials:
            tutorial.completed = False

    context = {
        'tutorials': tutorials,
        'student': student,
    }
    return render(request, 'management/tutorial_list.html', context)

@login_required
def student_tutorial_detail(request, pk):
    """Show a single tutorial and allow marking as complete."""
    tutorial = get_object_or_404(Tutorial, pk=pk, is_published=True)

    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        student = None

    if student:
        progress, created = StudentTutorialProgress.objects.get_or_create(
            student=student,
            tutorial=tutorial,
            defaults={'completed': False}
        )
    else:
        progress = None

    if request.method == 'POST' and student:
        if 'mark_complete' in request.POST:
            progress.completed = True
            progress.completed_at = timezone.now()
            progress.save()
            messages.success(request, 'Tutorial marked as complete!')
            return redirect('tutorial_detail', pk=tutorial.pk)
        elif 'mark_incomplete' in request.POST:
            progress.completed = False
            progress.completed_at = None
            progress.save()
            messages.success(request, 'Tutorial marked as incomplete.')
            return redirect('tutorial_detail', pk=tutorial.pk)

    context = {
        'tutorial': tutorial,
        'progress': progress,
        'student': student,
    }
    return render(request, 'management/tutorial_detail.html', context)      

class TutorialListView(LoginRequiredMixin, ListView):
    model = Tutorial
    template_name = 'management/tutorial_list_admin.html'
    context_object_name = 'tutorials'
    paginate_by = 20

    def get_queryset(self):
        # Teachers see all tutorials, including drafts
        return Tutorial.objects.all().select_related('created_by').order_by('-created_at')

class TutorialCreateView(LoginRequiredMixin,CreateView):
    model = Tutorial
    form_class = TutorialForm
    template_name = 'management/tutorial_form.html'
    success_url = reverse_lazy('tutorial_list_admin')

    def test_func(self):
        # Only staff or teachers can create
        return self.request.user.is_staff or self.request.user.groups.filter(name='Teacher').exists()

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

class TutorialUpdateView(LoginRequiredMixin, UpdateView):
    model = Tutorial
    form_class = TutorialForm
    template_name = 'management/tutorial_form.html'
    success_url = reverse_lazy('tutorial_list_admin')

    def test_func(self):
        # Only creator, staff, or teachers can edit
        tutorial = self.get_object()
        return (self.request.user.is_staff or 
                self.request.user.groups.filter(name='Teacher').exists() or
                tutorial.created_by == self.request.user)

class TutorialDeleteView(LoginRequiredMixin, DeleteView):
    model = Tutorial
    template_name = 'management/tutorial_confirm_delete.html'
    success_url = reverse_lazy('tutorial_list_admin')

    def test_func(self):
        tutorial = self.get_object()
        return (self.request.user.is_staff or 
                self.request.user.groups.filter(name='Teacher').exists() or
                tutorial.created_by == self.request.user) 
        
class TutorialProgressView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Shows all active students and their completion status for a specific tutorial.
    """
    model = Student
    template_name = 'management/tutorial_progress.html'
    context_object_name = 'students'
    paginate_by = 50

    def test_func(self):
        # Only staff or users in 'Teacher' group can view progress
        return self.request.user.is_staff or self.request.user.groups.filter(name='Teacher').exists()

    def get_queryset(self):
        # Store the tutorial for later use in context
        self.tutorial = get_object_or_404(Tutorial, pk=self.kwargs['pk'])
        # Return all active students (adjust if you need class filtering)
        return Student.objects.filter(is_active=True).order_by('Last_Name', 'First_Name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Fetch progress records for this tutorial
        progress_qs = StudentTutorialProgress.objects.filter(tutorial=self.tutorial)
        # Create a dictionary: student_id -> progress object
        progress_dict = {p.student_id: p for p in progress_qs}
        context['progress_dict'] = progress_dict
        context['tutorial'] = self.tutorial
        return context
    
@login_required
def after_login_redirect(request):
    """Redirect students to tutorials, others to dashboard."""
    if Student.objects.filter(user=request.user).exists():
        return redirect('tutorial_list')
    else:
        return redirect('dashboard')
    
class ClassCreateView(LoginRequiredMixin, CreateView):
    model = Class
    fields = ['name', 'grade_level', 'academic_year', 'teacher']
    template_name = 'management/class_form.html'
    success_url = reverse_lazy('class_list')
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.groups.filter(name='Admin').exists()  
    
    # Grade Views
class GradeListView(LoginRequiredMixin, ListView):
    model = Grade
    template_name = 'management/grade_list.html'
    context_object_name = 'grades'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Grade.objects.all().select_related('student', 'class_assigned', 'recorded_by')
        
        # Filter by student
        student_filter = self.request.GET.get('student')
        if student_filter:
            queryset = queryset.filter(student__Admission_Number=student_filter)
        
        # Filter by class
        class_filter = self.request.GET.get('class')
        if class_filter:
            queryset = queryset.filter(class_assigned_id=class_filter)
        
        # Filter by subject
        subject_filter = self.request.GET.get('subject')
        if subject_filter:
            queryset = queryset.filter(subject__icontains=subject_filter)
        
        # Filter by exam type
        exam_filter = self.request.GET.get('exam_type')
        if exam_filter:
            queryset = queryset.filter(exam_type=exam_filter)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        
        # Get the filtered queryset (without pagination for statistics)
        queryset = self.get_queryset()
        
        # Calculate statistics
        context['total_grades'] = queryset.count()
        
        # Calculate average grade (only if there are grades)
        avg_result = queryset.aggregate(avg_grade=Avg('grade'))
        context['avg_grade'] = round(avg_result['avg_grade'], 1) if avg_result['avg_grade'] else 0
        
        # Count unique students graded
        context['students_graded'] = queryset.values('student').distinct().count()
        
        # Count unique subjects
        context['total_subjects'] = queryset.values('subject').distinct().count()
        
        # For filter dropdowns
        context['students'] = Student.objects.filter(is_active=True)
        context['classes'] = Class.objects.all()
        context['exam_types'] = ['Midterm', 'Final', 'Quiz', 'Assignment']
        
        return context

class GradeCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Grade
    form_class = GradeForm
    template_name = 'management/grade_form.html'
    success_url = reverse_lazy('grade_list')
    
    def form_valid(self, form):
        form.instance.recorded_by = self.request.user
        messages.success(self.request, 'Grade recorded successfully!')
        return super().form_valid(form)
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.groups.filter(name='Teacher').exists()

class GradeUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Grade
    form_class = GradeForm
    template_name = 'management/grade_form.html'
    success_url = reverse_lazy('grade_list')
    
    def test_func(self):
        grade = self.get_object()
        return (self.request.user.is_staff or 
                self.request.user.groups.filter(name='Teacher').exists() or
                grade.recorded_by == self.request.user)

class GradeDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Grade
    success_url = reverse_lazy('grade_list')
    template_name = 'management/grade_confirm_delete.html'
    
    def test_func(self):
        grade = self.get_object()
        return (self.request.user.is_staff or 
                self.request.user.groups.filter(name='Teacher').exists() or
                grade.recorded_by == self.request.user)

class GradeDetailView(LoginRequiredMixin, DetailView):
    model = Grade
    template_name = 'management/grade_detail.html'
    context_object_name = 'grade'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grade = self.get_object()
        
        # Get other grades for the same student
        context['other_grades'] = Grade.objects.filter(
            student=grade.student
        ).exclude(id=grade.id).order_by('-recorded_at')[:5]
        
        # Get class average for this subject
        class_grades = Grade.objects.filter(
            class_assigned=grade.class_assigned,
            subject=grade.subject
        ).aggregate(
            avg_grade=Avg('grade'),
            count=Count('id')
        )
        context['class_average'] = round(class_grades['avg_grade'], 1) if class_grades['avg_grade'] else 0
        context['class_count'] = class_grades['count']
        
        return context

@login_required
def bulk_grade_entry(request):
    if request.method == 'POST':
        form = BulkGradeForm(request.POST)
        if form.is_valid():
            class_assigned = form.cleaned_data['class_assigned']
            subject = form.cleaned_data['subject']
            exam_type = form.cleaned_data['exam_type']
            
            students = Student.objects.filter(
                enrollment__class_enrolled=class_assigned,
                is_active=True
            )
            
            saved_count = 0
            updated_count = 0
            
            for student in students:
                grade_value = request.POST.get(f'grade_{student.id}')
                comments = request.POST.get(f'comments_{student.id}', '')
                
                if grade_value:
                    try:
                        grade = float(grade_value)
                        grade_obj, created = Grade.objects.update_or_create(
                            student=student,
                            class_assigned=class_assigned,
                            subject=subject,
                            defaults={
                                'grade': grade,
                                'exam_type': exam_type,
                                'comments': comments,
                                'recorded_by': request.user,
                            }
                        )
                        if created:
                            saved_count += 1
                        else:
                            updated_count += 1
                    except ValueError:
                        pass
            
            messages.success(request, f'Grades saved: {saved_count} new, {updated_count} updated!')
            return redirect('grade_list')
    else:
        form = BulkGradeForm()
    
    # Get students for the selected class
    class_id = request.GET.get('class')
    subject = request.GET.get('subject', '')
    exam_type = request.GET.get('exam_type', '')
    
    students = []
    existing_grades = {}
    existing_comments = {}
    
    if class_id and subject and exam_type:
        students = Student.objects.filter(
            enrollment__class_enrolled_id=class_id,
            is_active=True
        ).order_by('Last_Name', 'First_Name').distinct()
        
        # Load existing grades for editing
        for student in students:
            try:
                grade = Grade.objects.get(
                    student=student,
                    class_assigned_id=class_id,
                    subject=subject
                )
                existing_grades[student.id] = grade.grade
                existing_comments[student.id] = grade.comments
            except Grade.DoesNotExist:
                pass
    
    context = {
        'form': form,
        'students': students,
        'selected_class': class_id,
        'existing_grades': existing_grades,
        'existing_comments': existing_comments,
    }
    return render(request, 'management/grade_bulk_form.html', context)
@login_required
def student_transcript(request, pk):
    """Generate a transcript for a specific student"""
    student = get_object_or_404(Student, pk=pk)
    
    # Get all grades for this student
    grades = Grade.objects.filter(student=student).select_related('class_assigned')
    
    # Group by subject
    subjects = {}
    for grade in grades:
        if grade.subject not in subjects:
            subjects[grade.subject] = []
        subjects[grade.subject].append(grade)
    
    # Calculate average per subject
    subject_averages = {}
    for subject, grade_list in subjects.items():
        avg = sum(g.grade for g in grade_list) / len(grade_list)
        subject_averages[subject] = round(avg, 1)
    
    # Overall average
    all_grades = [g.grade for g in grades]
    overall_average = round(sum(all_grades) / len(all_grades), 1) if all_grades else 0
    
    context = {
        'student': student,
        'subjects': subjects,
        'subject_averages': subject_averages,
        'overall_average': overall_average,
        'grades': grades,
    }
    return render(request, 'management/student_transcript.html', context) 

@login_required
def export_attendance_excel(request):
    """Export attendance records as Excel file"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get date parameter
    selected_date_str = request.GET.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = date.today()
    else:
        selected_date = date.today()
    
    # Get attendance records for the selected date
    attendance_records = Attendance.objects.filter(
        attendance_date=selected_date
    ).select_related('student')
    
    # Headers
    headers = ['#', 'Admission Number', 'Student Name', 'Status', 'Check In Time', 'Check Out Time', 'Notes']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, record in enumerate(attendance_records, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=record.student.Admission_Number)
        ws.cell(row=row_idx, column=3, value=record.student.full_name)
        ws.cell(row=row_idx, column=4, value=record.get_status_display())
        ws.cell(row=row_idx, column=5, value=record.check_in_time.strftime('%H:%M') if record.check_in_time else '')
        ws.cell(row=row_idx, column=6, value=record.check_out_time.strftime('%H:%M') if record.check_out_time else '')
        ws.cell(row=row_idx, column=7, value=record.notes or '')
        
        # Apply borders to all cells in the row
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border
        
        # Color code status column
        status_cell = ws.cell(row=row_idx, column=4)
        if record.status == 'Present':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif record.status == 'Absent':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif record.status == 'Late':
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, 8):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary information at the bottom
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"Total Students: {attendance_records.count()}")
    ws.cell(row=summary_row + 1, column=2, value=f"Present: {attendance_records.filter(status='Present').count()}")
    ws.cell(row=summary_row + 2, column=2, value=f"Absent: {attendance_records.filter(status='Absent').count()}")
    ws.cell(row=summary_row + 3, column=2, value=f"Late: {attendance_records.filter(status='Late').count()}")
    ws.cell(row=summary_row + 4, column=2, value=f"Excused: {attendance_records.filter(status='Excused').count()}")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_{selected_date.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_attendance_report_excel(request):
    """Export attendance report as professional Excel file"""
    
    # Get date parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Set default dates (last 30 days if no dates provided)
    if not start_date_str:
        start_date = date.today() - timedelta(days=30)
    else:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = date.today() - timedelta(days=30)
    
    if not end_date_str:
        end_date = date.today()
    else:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = date.today()
    
    # Calculate statistics
    students = Student.objects.filter(is_active=True)
    attendance_records = Attendance.objects.filter(
        attendance_date__range=[start_date, end_date]
    )
    
    attendance_data = []
    total_present_all = 0
    total_absent_all = 0
    
    for student in students:
        student_records = attendance_records.filter(student=student)
        total_days = student_records.count()
        present_days = student_records.filter(status='Present').count()
        absent_days = student_records.filter(status='Absent').count()
        late_days = student_records.filter(status='Late').count()
        excused_days = student_records.filter(status='Excused').count()
        
        attendance_percentage = round((present_days / total_days * 100), 1) if total_days > 0 else 0
        
        attendance_data.append({
            'first_name': student.First_Name,
            'last_name': student.Last_Name,
            'admission_number': student.Admission_Number,
            'total_days': total_days,
            'present': present_days,
            'absent': absent_days,
            'late': late_days,
            'excused': excused_days,
            'percentage': attendance_percentage,
        })
        
        total_present_all += present_days
        total_absent_all += absent_days
    
    total_present_absent = total_present_all + total_absent_all
    overall_attendance = round((total_present_all / total_present_absent * 100), 1) if total_present_absent > 0 else 0
    total_days_range = (end_date - start_date).days + 1
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ========== SUMMARY SHEET ==========
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    accent_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_summary.merge_cells('A1:D1')
    cell = ws_summary['A1']
    cell.value = f"Attendance Report: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 30
    
    # Summary Statistics
    summary_data = [
        ['Metric', 'Value', '', ''],
        ['Report Period', f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}", '', ''],
        ['Total Days in Period', total_days_range, '', ''],
        ['Total Students', students.count(), '', ''],
        ['Total Present Days', total_present_all, '', ''],
        ['Total Absent Days', total_absent_all, '', ''],
        ['Overall Attendance Rate', f"{overall_attendance}%", '', ''],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
            ws_summary.column_dimensions[chr(64 + col_idx)].width = 25
    
    # ========== DETAILS SHEET ==========
    ws_details = wb.create_sheet("Student Details")
    
    # Headers
    headers = [
        '#',
        'Admission Number',
        'Student Name',
        'Total Days',
        'Present',
        'Absent',
        'Late',
        'Excused',
        'Attendance %',
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_idx, student in enumerate(attendance_data, 2):
        # Serial number
        cell = ws_details.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Admission Number
        cell = ws_details.cell(row=row_idx, column=2, value=student['admission_number'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Student Name
        cell = ws_details.cell(row=row_idx, column=3, value=f"{student['first_name']} {student['last_name']}")
        cell.border = border
        
        # Total Days
        cell = ws_details.cell(row=row_idx, column=4, value=student['total_days'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Present (Green)
        cell = ws_details.cell(row=row_idx, column=5, value=student['present'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if student['present'] > 0:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Absent (Red)
        cell = ws_details.cell(row=row_idx, column=6, value=student['absent'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if student['absent'] > 0:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Late (Yellow)
        cell = ws_details.cell(row=row_idx, column=7, value=student['late'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if student['late'] > 0:
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Excused
        cell = ws_details.cell(row=row_idx, column=8, value=student['excused'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Attendance Percentage
        cell = ws_details.cell(row=row_idx, column=9, value=student['percentage'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0"%'
        
        # Color code percentage
        if student['percentage'] >= 80:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif student['percentage'] >= 60:
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, 10):
        max_length = 0
        for row in range(1, ws_details.max_row + 1):
            cell_value = ws_details.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 30)
        ws_details.column_dimensions[chr(64 + col)].width = adjusted_width
    
    # Add total row
    total_row = ws_details.max_row + 2
    ws_details.cell(row=total_row, column=3, value="TOTAL:").font = Font(bold=True)
    ws_details.cell(row=total_row, column=4, value=f"=SUM(D2:D{ws_details.max_row})").font = Font(bold=True)
    ws_details.cell(row=total_row, column=5, value=f"=SUM(E2:E{ws_details.max_row})").font = Font(bold=True)
    ws_details.cell(row=total_row, column=6, value=f"=SUM(F2:F{ws_details.max_row})").font = Font(bold=True)
    ws_details.cell(row=total_row, column=7, value=f"=SUM(G2:G{ws_details.max_row})").font = Font(bold=True)
    ws_details.cell(row=total_row, column=8, value=f"=SUM(H2:H{ws_details.max_row})").font = Font(bold=True)
    
    # ========== CHART SHEET (Optional) ==========
    ws_chart = wb.create_sheet("Attendance Chart")
    
    # Create a bar chart
    chart = BarChart()
    chart.title = "Attendance by Student"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    
    # Data for chart (top 10 students)
    top_students = sorted(attendance_data, key=lambda x: x['percentage'], reverse=True)[:10]
    
    for idx, student in enumerate(top_students, 1):
        ws_chart.cell(row=idx + 1, column=1, value=f"{student['first_name']} {student['last_name']}")
        ws_chart.cell(row=idx + 1, column=2, value=student['percentage'])
    
    ws_chart.cell(row=1, column=1, value="Student Name")
    ws_chart.cell(row=1, column=2, value="Attendance %")
    
    data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(top_students) + 1, max_col=2)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(top_students) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_chart.add_chart(chart, "E5")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def export_fees_excel(request):
    """Export fee records as professional Excel file"""
    
    # Get filtered queryset
    queryset = Fee.objects.select_related('student', 'fee_type').all()
    
    # Apply filters
    status_filter = request.GET.get('status')
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    student_filter = request.GET.get('student')
    if student_filter:
        queryset = queryset.filter(student__Admission_Number=student_filter)
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            queryset = queryset.filter(due_date__gte=start_date)
        except ValueError:
            pass
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(due_date__lte=end_date)
        except ValueError:
            pass
    
    # Calculate summary statistics
    total_due = queryset.aggregate(total=Sum('amount_due'))['total'] or 0
    total_paid = queryset.aggregate(total=Sum('amount_paid'))['total'] or 0
    total_balance = total_due - total_paid
    paid_count = queryset.filter(status='Paid').count()
    unpaid_count = queryset.filter(status='Unpaid').count()
    partial_count = queryset.filter(status='partially paid').count()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== SUMMARY SHEET ==========
    ws_summary = wb.active
    ws_summary.title = "Fee Summary"
    
    # ✅ FIXED: Set values BEFORE merging
    ws_summary['A1'] = "FEE COLLECTION REPORT"
    ws_summary['A2'] = f"Generated on: {date.today().strftime('%B %d, %Y at %H:%M')}"
    
    # Now merge AFTER setting values
    ws_summary.merge_cells('A1:D1')
    ws_summary.merge_cells('A2:D2')
    
    # Apply styles to merged cells
    cell_a1 = ws_summary['A1']
    cell_a1.font = title_font
    cell_a1.fill = title_fill
    cell_a1.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 35
    
    cell_a2 = ws_summary['A2']
    cell_a2.alignment = Alignment(horizontal='center')
    
    # Summary Statistics
    summary_data = [
        ['SUMMARY STATISTICS', '', '', ''],
        ['Total Records', queryset.count(), 'Total Amount Due', f"Kshs {total_due:,.2f}"],
        ['Paid Records', paid_count, 'Total Amount Paid', f"Kshs {total_paid:,.2f}"],
        ['Unpaid Records', unpaid_count, 'Total Outstanding', f"Kshs {total_balance:,.2f}"],
        ['Partially Paid', partial_count, 'Collection Rate', f"{round((total_paid/total_due)*100,1) if total_due > 0 else 0}%"],
    ]
    
    start_row = 4
    for row_idx, row_data in enumerate(summary_data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == start_row:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            if col_idx == 1 or col_idx == 3:
                cell.font = Font(bold=True)
    
    # Adjust column widths
    for col in range(1, 5):
        ws_summary.column_dimensions[chr(64 + col)].width = 25
    
    # ========== DETAILS SHEET ==========
    ws_details = wb.create_sheet("Fee Details")
    
    # Headers
    headers = ['#', 'Student Name', 'Admission No', 'Fee Type', 'Due Date', 'Amount Due', 'Amount Paid', 'Balance', 'Status', 'Payment Method', 'Transaction ID', 'Date Paid']
    
    for col, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_idx, fee in enumerate(queryset, 2):
        balance = fee.amount_due - fee.amount_paid
        
        # Serial number
        cell = ws_details.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Student Name
        cell = ws_details.cell(row=row_idx, column=2, value=fee.student.full_name)
        cell.border = border
        
        # Admission Number
        cell = ws_details.cell(row=row_idx, column=3, value=fee.student.Admission_Number)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Fee Type
        cell = ws_details.cell(row=row_idx, column=4, value=fee.fee_type.name if fee.fee_type else '—')
        cell.border = border
        
        # Due Date
        cell = ws_details.cell(row=row_idx, column=5, value=fee.due_date.strftime('%d/%m/%Y') if fee.due_date else '')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Amount Due
        cell = ws_details.cell(row=row_idx, column=6, value=float(fee.amount_due))
        cell.border = border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        
        # Amount Paid
        cell = ws_details.cell(row=row_idx, column=7, value=float(fee.amount_paid))
        cell.border = border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        
        # Balance
        cell = ws_details.cell(row=row_idx, column=8, value=float(balance))
        cell.border = border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        if balance > 0:
            cell.font = Font(color="FF0000")
        
        # Status with color
        cell = ws_details.cell(row=row_idx, column=9, value=fee.get_status_display())
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if fee.status == 'Paid':
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif fee.status == 'Unpaid':
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif fee.status == 'partially paid':
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Payment Method
        cell = ws_details.cell(row=row_idx, column=10, value=fee.payment_method or '—')
        cell.border = border
        
        # Transaction ID
        cell = ws_details.cell(row=row_idx, column=11, value=fee.transaction_id or '—')
        cell.border = border
        
        # Date Paid
        cell = ws_details.cell(row=row_idx, column=12, value=fee.paid_date.strftime('%d/%m/%Y') if fee.paid_date else '—')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col in range(1, 13):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(1, ws_details.max_row + 1):
            cell_value = ws_details.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 30)
        ws_details.column_dimensions[col_letter].width = adjusted_width
    
    # Add total row
    total_row = ws_details.max_row + 2
    ws_details.cell(row=total_row, column=5, value="TOTAL:").font = Font(bold=True)
    ws_details.cell(row=total_row, column=6, value=f"=SUM(F2:F{ws_details.max_row})").font = Font(bold=True)
    ws_details.cell(row=total_row, column=7, value=f"=SUM(G2:G{ws_details.max_row})").font = Font(bold=True)
    ws_details.cell(row=total_row, column=8, value=f"=SUM(H2:H{ws_details.max_row})").font = Font(bold=True)
    
    # ========== CHART SHEET ==========
    ws_chart = wb.create_sheet("Payment Analysis")
    
    # Payment by Status Pie Chart
    status_data = [
        ['Status', 'Count'],
        ['Paid', paid_count],
        ['Unpaid', unpaid_count],
        ['Partially Paid', partial_count],
    ]
    
    for row_idx, row_data in enumerate(status_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws_chart.cell(row=row_idx, column=col_idx, value=value)
    
    # Create pie chart
    pie_chart = PieChart()
    pie_chart.title = "Payment Status Distribution"
    pie_chart.width = 15
    pie_chart.height = 10
    
    data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(status_data), max_col=2)
    labels = Reference(ws_chart, min_col=1, min_row=2, max_row=len(status_data))
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    ws_chart.add_chart(pie_chart, "D5")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"fee_report_{date.today().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# Timetable Views
class TimetableEntryListView(LoginRequiredMixin, ListView):
    model = TimetableEntry
    template_name = 'management/timetable_list.html'
    context_object_name = 'entries'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = TimetableEntry.objects.all().select_related('class_assigned', 'day', 'period', 'teacher')
        
        class_filter = self.request.GET.get('class')
        if class_filter:
            queryset = queryset.filter(class_assigned_id=class_filter)
        
        day_filter = self.request.GET.get('day')
        if day_filter:
            queryset = queryset.filter(day_id=day_filter)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['classes'] = Class.objects.all()
        context['days'] = Day.objects.filter(is_active=True)
        return context

class TimetableEntryCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = TimetableEntry
    form_class = TimetableEntryForm
    template_name = 'management/timetable_form.html'
    success_url = reverse_lazy('timetable_list')
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.groups.filter(name='Teacher').exists()

class TimetableEntryUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = TimetableEntry
    form_class = TimetableEntryForm
    template_name = 'management/timetable_form.html'
    success_url = reverse_lazy('timetable_list')
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.groups.filter(name='Teacher').exists()

class TimetableEntryDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = TimetableEntry
    success_url = reverse_lazy('timetable_list')
    template_name = 'management/timetable_confirm_delete.html'
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.groups.filter(name='Teacher').exists()

@login_required
def view_class_timetable(request, class_id):
    """View timetable for a specific class in grid format"""
    class_obj = get_object_or_404(Class, pk=class_id)
    days = Day.objects.filter(is_active=True)
    periods = Period.objects.all()
    
    # Build timetable grid
    timetable_grid = {}
    for day in days:
        timetable_grid[day.id] = {}
        for period in periods:
            entry = TimetableEntry.objects.filter(
                class_assigned=class_obj,
                day=day,
                period=period
            ).first()
            timetable_grid[day.id][period.id] = entry
    
    context = {
        'class_obj': class_obj,
        'days': days,
        'periods': periods,
        'timetable_grid': timetable_grid,
    }
    return render(request, 'management/timetable_view.html', context)

@login_required
def copy_timetable(request):
    """Copy timetable from one class to another"""
    if request.method == 'POST':
        form = BulkTimetableForm(request.POST)
        if form.is_valid():
            source_class = form.cleaned_data['source_class']
            target_class = form.cleaned_data['target_class']
            academic_year = form.cleaned_data['academic_year']
            term = form.cleaned_data['term']
            
            # Delete existing entries for target class
            TimetableEntry.objects.filter(
                class_assigned=target_class,
                academic_year=academic_year,
                term=term
            ).delete()
            
            # Copy entries from source class
            source_entries = TimetableEntry.objects.filter(
                class_assigned=source_class,
                academic_year=academic_year,
                term=term
            )
            
            copied_count = 0
            for entry in source_entries:
                TimetableEntry.objects.create(
                    class_assigned=target_class,
                    day=entry.day,
                    period=entry.period,
                    subject=entry.subject,
                    teacher=entry.teacher,
                    room=entry.room,
                    is_break=entry.is_break,
                    break_name=entry.break_name,
                    academic_year=academic_year,
                    term=term
                )
                copied_count += 1
            
            messages.success(request, f'Successfully copied {copied_count} timetable entries from {source_class.name} to {target_class.name}')
            return redirect('timetable_list')
    else:
        form = BulkTimetableForm()
    
    return render(request, 'management/timetable_copy.html', {'form': form})

@login_required
def my_timetable(request):
    """Show timetable for the logged-in teacher"""
    if not request.user.groups.filter(name='Teacher').exists():
        messages.error(request, 'Access denied. Only teachers can view this page.')
        return redirect('dashboard')
    
    days = Day.objects.filter(is_active=True)
    periods = Period.objects.all()
    
    # Build timetable grid for teacher
    timetable_grid = {}
    for day in days:
        timetable_grid[day.id] = {}
        for period in periods:
            entry = TimetableEntry.objects.filter(
                teacher=request.user,
                day=day,
                period=period
            ).first()
            timetable_grid[day.id][period.id] = entry
    
    context = {
        'days': days,
        'periods': periods,
        'timetable_grid': timetable_grid,
        'is_teacher_view': True,
    }
    return render(request, 'management/timetable_teacher.html', context)

class PeriodListView(LoginRequiredMixin, ListView):
    model = Period
    template_name = 'management/period_list.html'
    context_object_name = 'periods'

class PeriodCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Period
    form_class = PeriodForm
    template_name = 'management/period_form.html'
    success_url = reverse_lazy('period_list')
    
    def test_func(self):
        return self.request.user.is_staff
    
class PeriodUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Period
    form_class = PeriodForm
    template_name = 'management/period_form.html'
    success_url = reverse_lazy('period_list')
    
    def test_func(self):
        return self.request.user.is_staff

class PeriodDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Period
    success_url = reverse_lazy('period_list')
    template_name = 'management/period_confirm_delete.html'
    
    def test_func(self):
        return self.request.user.is_staff

class DayListView(LoginRequiredMixin, ListView):
    model = Day
    template_name = 'management/day_list.html'
    context_object_name = 'days'

class DayCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Day
    form_class = DayForm
    template_name = 'management/day_form.html'
    success_url = reverse_lazy('day_list')
    
    def test_func(self):
        return self.request.user.is_staff
    
class DayUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Day
    form_class = DayForm
    template_name = 'management/day_form.html'
    success_url = reverse_lazy('day_list')
    
    def test_func(self):
        return self.request.user.is_staff

class DayDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Day
    success_url = reverse_lazy('day_list')
    template_name = 'management/day_confirm_delete.html'
    
    def test_func(self):
        return self.request.user.is_staff